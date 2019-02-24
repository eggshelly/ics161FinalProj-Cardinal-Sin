#DialogueConverter.py
# -*- coding: utf-8 -*-

#Michaela Gonzales
#michacg2@uci.edu

"""
This file reads from a word document and creates a an excel file.
Will appropriately place dialogue lines into cell to use for UNITY. 

ASSUMPTIONS:
    -Dialogue.txt is in the same directory as this file.

"""
##HEADER INITIALIZATION##
from openpyxl import Workbook, load_workbook
import csv

##VARIABLES##
FILENAME = "Dialogue.txt"

##SCRIPT##

def run(file):   
    canmakewb = False
    text = list()

    for line in file:
        if line == "INTRODUCTION\n" or (" " in line and line[:line.index(" ")] in ["INTRODUCTION", "WINTER", "SPRING", "SUMMER", "AUTUMN"]): #means new wb!
            print(line)
            if canmakewb:
                makewb(title, text)
            canmakewb = True

            title = line.rstrip()
            text  = list()

        else:
            text.append(line.rstrip())
            
    makewb(title,text)


def makewb(title:str, text:str):
    wb = Workbook()
    ws = wb.active
    ws.cell(1,1).value = "SPEAKER"
    ws.cell(1,2).value = "CONVO"
    ws.cell(1,3).value = "SPRITE"

    for line in text:
        row = ws.max_row + 1
        ws.cell(row, 1).value  = line[:line.index(":")]
        ws.cell(row, 2).value  = line[line.index(":")+1:]
        ws.cell(row, 3).value  = "default"

    wb.save(f"{title}.csv")


def cleanup(file):
    file.close()
    
if __name__ == "__main__":
    file = open(FILENAME, "r")
    run(file)
    cleanup(file)
    print("DialogueConverter.py: OUTPUT END")
